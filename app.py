
import os
import re
import pandas as pd
from datetime import datetime
from docx import Document
from openpyxl import Workbook, load_workbook

# Define column headers for the output Excel
columns = [
    "Name of IRBn/Bn", "Reserves Deployed (Distt/Strength/Duration/In-Charge)",
    "Districts where force deployed", "Stay Arrangement/Bathrooms (Quality)",
    "Messing Arrangements", "CO's last Interaction with SP",
    "Disciplinary Issues", "Reserves Detained", "Training",
    "Welfare Initiative in Last 24 Hrs", "Reserves Available in Bn",
    "Issue for AP&T/PHQ"
]

# Initialize the workbook and worksheet
output_path = "/mnt/data/IRBn_Daily_Report_Compiled.xlsx"
if not os.path.exists(output_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consolidated Report"
    sheet.append(columns)
    workbook.save(output_path)
else:
    workbook = load_workbook(output_path)
    sheet = workbook.active

# Directory containing Word files
input_dir = "/mnt/data"
files = [f for f in os.listdir(input_dir) if f.endswith(".docx")]

for file in files:
    doc_path = os.path.join(input_dir, file)
    doc = Document(doc_path)

    full_text = "\n".join([para.text.strip() for para in doc.paragraphs if para.text.strip()])

    def extract(pattern, group=1, default="Nil"):
        match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
        return match.group(group).strip() if match else default

    bn_name = extract(r"Bn\s*:?\s*(?:No\.\s*)?(.*?)(?=Name of CO|Name & Rank|Name of C.O|1\.|\n)", default="Unknown Bn").replace("Bn", "Bn").strip(":- ")
    deployed = extract(r"1[\.\)]\s*Detail.*?Reserves.*?:\s*(.*?)(?=2[\.\)]|Stay arrangements|\n3[\.\)])", default="Nil")
    deployed_cleaned = " ".join(deployed.split())
    districts = ", ".join(sorted(set(re.findall(r"(Shimla|Kullu|Mandi|Solan|Sirmaur|Bilaspur|Kangra|Nahan|Chamba|Hamirpur|Una|Lahaul|Spiti|Kinnaur|Baddi|Dharamshala|D/Shala|Nagrota|Pandoh|Jangalberi|Bassi|Sakoh|Kotkhai|Nirmand|Thunag|Kaithu|Kanda)", deployed, re.IGNORECASE))))
    stay = extract(r"2[\.\)]\s*Stay arrangements.*?:\s*(.*?)(?=3[\.\)]|\n4[\.\)])", default="Nil").replace("\n", " ")
    mess = extract(r"3[\.\)]\s*Messing arrangements.*?:\s*(.*?)(?=4[\.\)]|\n5[\.\)])", default="Nil").replace("\n", " ")
    interaction = extract(r"4[\.\)]\s*Date on which CO.*?\s*:?\s*(.*?)(?=5[\.\)]|\n6[\.\)])", default="Nil").replace("\n", " ")
    discipline = extract(r"5[\.\)]\s*Any (incident of )?disciplinary issue.*?:\s*(.*?)(?=6[\.\)]|\n7[\.\)])", default="Nil").replace("\n", " ")
    detained = extract(r"6[\.\)]\s*Reserves detained.*?:\s*(.*?)(?=7[\.\)]|\n8[\.\)])", default="Nil").replace("\n", " ")
    training = extract(r"7[\.\)]\s*Training.*?:\s*(.*?)(?=8[\.\)]|\n9[\.\)])", default="Nil").replace("\n", " ")
    welfare = extract(r"8[\.\)]\s*Any Initiative.*?:\s*(.*?)(?=9[\.\)]|\n10[\.\)])", default="Nil").replace("\n", " ")
    reserves = extract(r"9[\.\)]\s*Reserves.*?:\s*(.*?)(?=10[\.\)]|\n11[\.\)])", default="Nil").replace("\n", " ")
    issue = extract(r"10[\.\)]\s*Any other issue.*?:\s*(.*)", default="Nil").replace("\n", " ")

    sheet.append([
        bn_name, deployed_cleaned, districts, stay, mess,
        interaction, discipline, detained, training, welfare,
        reserves, issue
    ])

# Save the final workbook
workbook.save(output_path)
