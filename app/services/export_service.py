"""Export service - Enhanced Excel, CSV, and JSON generation."""

import io
from typing import Dict, List, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """Enhanced service for exporting reports in various formats."""
    
    FIELD_NAMES = [
        "Name of IRBn/Bn",
        "Reserves Deployed",
        "Districts where force deployed",
        "Stay Arrangement/Bathrooms",
        "Messing Arrangements",
        "CO's last Interaction with SP",
        "Disciplinary Issues",
        "Reserves Detained",
        "Training",
        "Welfare Initiative in Last 24 Hrs",
        "Reserves Available in Bn",
        "Issue for AP&T/PHQ"
    ]
    
    FIELD_KEYS = [
        "unit_name",
        "reserves_deployed", 
        "districts",
        "stay_arrangement",
        "messing",
        "co_interaction_date",
        "disciplinary_issues",
        "reserves_detained",
        "training",
        "welfare",
        "reserves_available",
        "issues_for_phq"
    ]
    
    # Excel color scheme
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HIGH_CONF_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    MED_CONF_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    LOW_CONF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def export_excel(self, report_data: Dict, confidence_scores: Dict, include_metadata: bool = True) -> bytes:
        """Generate enhanced Excel file with formatting."""
        output = io.BytesIO()
        workbook = Workbook()
        
        # === Sheet 1: Main Report ===
        ws_main = workbook.active
        ws_main.title = "HP Police Report"
        
        # Prepare data
        data_row = [report_data.get(key, "") for key in self.FIELD_KEYS]
        conf_row = [confidence_scores.get(key, 0) for key in self.FIELD_KEYS]
        
        # Header row
        headers = ["Field", "Value", "Confidence", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        
        # Data rows
        for row_idx, (field_name, field_key) in enumerate(zip(self.FIELD_NAMES, self.FIELD_KEYS), 2):
            # Field name
            cell = ws_main.cell(row=row_idx, column=1, value=field_name)
            cell.font = Font(bold=True)
            cell.border = self.BORDER
            
            # Value
            value = report_data.get(field_key, "")
            cell = ws_main.cell(row=row_idx, column=2, value=value)
            cell.border = self.BORDER
            cell.alignment = Alignment(wrap_text=True)
            
            # Confidence score
            confidence = confidence_scores.get(field_key, 0)
            cell = ws_main.cell(row=row_idx, column=3, value=f"{confidence:.1%}")
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center')
            
            # Status
            if confidence >= 0.7:
                status = "High"
                fill = self.HIGH_CONF_FILL
            elif confidence >= 0.5:
                status = "Medium"
                fill = self.MED_CONF_FILL
            else:
                status = "Low"
                fill = self.LOW_CONF_FILL
            
            cell = ws_main.cell(row=row_idx, column=4, value=status)
            cell.fill = fill
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws_main.column_dimensions['A'].width = 30
        ws_main.column_dimensions['B'].width = 50
        ws_main.column_dimensions['C'].width = 12
        ws_main.column_dimensions['D'].width = 12
        
        # Freeze header row
        ws_main.freeze_panes = 'A2'
        
        # === Sheet 2: Summary Statistics ===
        ws_summary = workbook.create_sheet("Summary")
        
        # Statistics
        stats = [
            ["Total Fields", len(self.FIELD_KEYS)],
            ["Fields with Data", sum(1 for v in report_data.values() if v and str(v).strip())],
            ["High Confidence Fields", sum(1 for c in confidence_scores.values() if c >= 0.7)],
            ["Medium Confidence Fields", sum(1 for c in confidence_scores.values() if 0.5 <= c < 0.7)],
            ["Low Confidence Fields", sum(1 for c in confidence_scores.values() if c < 0.5)],
            ["Average Confidence", f"{sum(confidence_scores.values()) / len(confidence_scores):.1%}"],
            ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        
        for row_idx, (label, value) in enumerate(stats, 1):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # === Sheet 3: Confidence Details (optional) ===
        if include_metadata:
            ws_conf = workbook.create_sheet("Confidence Details")
            
            for col, field_name in enumerate(self.FIELD_NAMES, 1):
                cell = ws_conf.cell(row=1, column=col, value=field_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            
            for col, field_key in enumerate(self.FIELD_KEYS, 1):
                confidence = confidence_scores.get(field_key, 0)
                cell = ws_conf.cell(row=2, column=col, value=confidence)
                
                # Color code based on confidence
                if confidence >= 0.7:
                    cell.fill = self.HIGH_CONF_FILL
                elif confidence >= 0.5:
                    cell.fill = self.MED_CONF_FILL
                else:
                    cell.fill = self.LOW_CONF_FILL
            
            # Make headers readable
            for col in range(1, len(self.FIELD_NAMES) + 1):
                ws_conf.column_dimensions[get_column_letter(col)].width = 25
        
        # Write and return
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_excel_multi(self, reports: List[Dict], confidence_scores_list: List[Dict]) -> bytes:
        """Export multiple reports to Excel."""
        output = io.BytesIO()
        workbook = Workbook()
        
        # Main data sheet
        ws = workbook.active
        ws.title = "All Reports"
        
        # Headers
        headers = ["Report #"] + self.FIELD_NAMES
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        # Data rows
        for report_idx, (report_data, conf_scores) in enumerate(zip(reports, confidence_scores_list), 1):
            ws.cell(row=report_idx + 1, column=1, value=f"Report {report_idx}")
            
            for col, field_key in enumerate(self.FIELD_KEYS, 2):
                value = report_data.get(field_key, "")
                ws.cell(row=report_idx + 1, column=col, value=value)
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        for col in range(2, len(self.FIELD_NAMES) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 30
        
        ws.freeze_panes = 'A2'
        
        # Write and return
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_csv(self, report_data: Dict) -> bytes:
        """Generate CSV file."""
        data = {name: report_data.get(key, "") for name, key in zip(self.FIELD_NAMES, self.FIELD_KEYS)}
        df = pd.DataFrame([data])
        
        output = io.StringIO()
        df.to_csv(output, index=False)
        
        return output.getvalue().encode('utf-8')
    
    def export_json(self, report_data: Dict, confidence_scores: Dict, metadata: Dict = None) -> str:
        """Generate JSON export with metadata."""
        import json
        
        export_data = {
            "report": {
                key: report_data.get(key, "") for key in self.FIELD_KEYS
            },
            "confidence_scores": {
                key: round(confidence_scores.get(key, 0), 3) for key in self.FIELD_KEYS
            },
            "statistics": {
                "total_fields": len(self.FIELD_KEYS),
                "fields_filled": sum(1 for v in report_data.values() if v and str(v).strip()),
                "average_confidence": round(sum(confidence_scores.values()) / len(confidence_scores), 3),
                "high_confidence_count": sum(1 for c in confidence_scores.values() if c >= 0.7),
                "medium_confidence_count": sum(1 for c in confidence_scores.values() if 0.5 <= c < 0.7),
                "low_confidence_count": sum(1 for c in confidence_scores.values() if c < 0.5),
            }
        }
        
        if metadata:
            export_data["metadata"] = metadata
        
        export_data["generated_at"] = datetime.now().isoformat()
        
        return json.dumps(export_data, indent=2, default=str)


# Singleton
export_service = ExportService()
